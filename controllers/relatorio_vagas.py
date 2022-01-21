from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import date


class Planilha:
    def __init__(self) -> None:
        # Cria uma pasta de trabalho para conter a planilha
        self.arquivo_excel = Workbook()
        # # Ativa a guia que já vem por padrão no Workbook
        self.sheetVagas = self.arquivo_excel.active
        # # Pega a data de hoje para nomear a guia da planilha
        self.today = date.today()

    def formatar_planilha(self) -> None:
        # Renomenado o título da guia para a data de hoje
        self.sheetVagas.title = format(self.today)

        # # Escrevendo no arquivo o Nome/Local e Descrição da vaga
        self.sheetVagas['A1'] = "Nome"
        self.sheetVagas['B1'] = "Local"
        self.sheetVagas['C1'] = "Descrição"

        self.sheetVagas['A1'].font = Font(
            size=18, bold=True)  # Fonte em Negrito
        self.sheetVagas['B1'].font = Font(
            size=18, bold=True)  # Fonte em Negrito
        self.sheetVagas['C1'].font = Font(
            size=18, bold=True)  # Fonte em Negrito

        self.sheetVagas['A1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sheetVagas['B1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sheetVagas['C1'].alignment = Alignment(
            horizontal='center', vertical='center')

        # Dimensão da coluna A
        self.sheetVagas.column_dimensions['A'].width = 50

        # Dimensão da coluna B
        self.sheetVagas.column_dimensions['B'].width = 20

        # Dimensão da coluna C
        self.sheetVagas.column_dimensions['C'].width = 70

        self.arquivo_excel.save('vagas.xlsx')

    def escreve_dados(self, lista_vagas: list, lista_locais: list, lista_descricao: list) -> None:

        for i in range(len(lista_vagas)):
            self.sheetVagas.cell(row=i+2, column=1).value = lista_vagas[i]
            self.sheetVagas.cell(row=i+2, column=2).value = lista_locais[i]
            self.sheetVagas.cell(row=i+2, column=3).value = lista_descricao[i]
            self.sheetVagas[f'A{i+2}'].alignment = Alignment(vertical='center')
            self.sheetVagas[f'B{i+2}'].alignment = Alignment(vertical='center')
            self.sheetVagas[f'C{i+2}'].alignment = Alignment(wrap_text=True)

        self.arquivo_excel.save('vagas.xlsx')
