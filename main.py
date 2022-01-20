from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import date
import re
from fileinput import filename
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Cria uma pasta de trabalho para conter a planilha
arquivo_excel = Workbook()

# # Ativa a guia que já vem por padrão no Workbook
sheetVagas = arquivo_excel.active

# # Pega a data de hoje para nomear a guia da planilha
today = date.today()

# # Renomenado o título da guia para a data de hoje
sheetVagas.title = format(today)

# # Escrevendo no arquivo o Nome/Local e Descrição da vaga
sheetVagas['A1'] = "Nome"
sheetVagas['B1'] = "Local"
sheetVagas['C1'] = "Descrição"

sheetVagas['A1'].font = Font(bold=True)  # Fonte em Negrito
sheetVagas['B1'].font = Font(bold=True)  # Fonte em Negrito
sheetVagas['C1'].font = Font(bold=True)  # Fonte em Negrito

sheetVagas.column_dimensions['A'].width = 50  # Dimensão da coluna A

sheetVagas.column_dimensions['B'].width = 20  # Dimensão da coluna B

sheetVagas.column_dimensions['C'].width = 70  # Dimensão da coluna C

#sheetVagas.row_dimensions[1].height = 100

#sheetVagas['C2'] = " O que você terá que fazer: Profissional irá atuar na área de Suporte Técnico à Usuário N2, manutenção preventiva e corretiva de notebooks e periféricos. Realizará a instalação e configuração de aplicativo, verificação hardware e instalação de sistema operacional. Análises de equipamentos de troca ou desligamento, substituição de processadores, fontes, hd, memórias. Controle de ativos que entram e saem do estoque. • Área e especialização profissional: Informática, TI, Telecomunicações - TI, Logistica, Administração • Nível hierárquico: Estagiário • Local de trabalho: São Judas - SP"

#sheetVagas.alignment = Alignment(wrap_text=True)

# Começando a parte de pegar os dados do site
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--lang=pt-BR')
chrome_options.add_argument('disable-infobars')
chrome_options.add_argument('--log-level=3')
driver = webdriver.Chrome(
    options=chrome_options)

# inicia o navegador no site da cadmus
driver.get('https://cadmus.com.br/vagas-tecnologia/')


time.sleep(5)

# busca todas as vagas

vagas = driver.find_elements(By.XPATH, '//div[@class="box"]')
time.sleep(2)


# descrição
dados_vagas = []

# Contador para controlar o número de linhas da planilha
i = 0

# coleta as vagas e os locais das vagas
driver.execute_script("window.scrollTo(0,1000);")
time.sleep(2)
for vaga in vagas:

    titulo_el = vaga.find_element(By.TAG_NAME, 'h3')
    local_el = vaga.find_element(By.CLASS_NAME, 'local')
    titulo = titulo_el.text
    local = local_el.text
    print(f'Vaga: {titulo}')
    print(f'Local: {local}')

    # Titulo é chave

    dados_vagas.append({titulo: {
        'local': local,
        'descricao': None
    }})

    # Inserindo título e local da linha 2 em diante e coluna A1 e B1 respectivamente
    sheetVagas.cell(row=i+2, column=1).value = titulo
    sheetVagas.cell(row=i+2, column=2).value = local
    i += 1

time.sleep(5)
i = 0


for vaga in dados_vagas:
    driver.execute_script("window.scrollTo(0,1000);")
    time.sleep(1)
    btn_descricao = '//div[@class="box"]/h3[contains(text(), "{}")]/../p/a'.format(
        list(vaga.keys())[0])
    link = driver.find_element(
        By.XPATH, btn_descricao)
    link.send_keys(Keys.RETURN)
    descricao = driver.find_element(
        By.XPATH, '//div[@class="box z-depth-1"]/p')
    print(descricao.text)

    # Retirar o '\n' do texto
    new = re.sub('\n', ' ', descricao.text)

    print(new)

    sheetVagas.cell(row=i+2, column=3).value = new
    i += 1
    driver.back()


# print(dados_vagas)

driver.close()

#  Salvando o arquivo

arquivo_excel.save('vagas.xlsx')

# Rementente
fromaddr = 'robovagas@gmail.com'
# Destinatários
toaddr = 'bruno.cabral@cadmus.com.br'

msg = MIMEMultipart()
msg['From'] = fromaddr
msg['To'] = toaddr

# Assunto do email
msg['Subject'] = "E-mail de test"
# Corpo do emial
body = "Email enviado do nosso robô"

msg.attach(MIMEText(body, 'plain'))

# Arquivo a ser anexado
filename = "vagas.xlsx"
anexo = open("vagas.xlsx", "rb")

p = MIMEBase('aplication', 'octet-stream')
p.set_payload((anexo).read())
encoders.encode_base64(p)
p.add_header("Content-Disposition", "attachment; filename= %s" % filename)
msg.attach(p)

s = smtplib.SMTP('smtp.gmail.com', 587)

# Segurança
s.starttls()

s.login(fromaddr, '123Batatinha')

# Converte para String
text = msg.as_string()

s.sendmail(fromaddr, toaddr, text)

s.quit()
